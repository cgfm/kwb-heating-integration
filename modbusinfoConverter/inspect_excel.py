#!/usr/bin/env python3
"""Inspect ModbusInfo Excel file structure."""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl library is required")
    sys.exit(1)

# Get the first Excel file
script_dir = Path(__file__).parent
input_dir = script_dir / "modbusinfo"
xlsx_files = list(input_dir.glob("ModbusInfo*.xlsx"))

if not xlsx_files:
    print(f"No Excel files found in {input_dir}")
    sys.exit(1)

xlsx_file = xlsx_files[0]
print(f"Inspecting: {xlsx_file.name}\n")

# Load workbook
workbook = openpyxl.load_workbook(xlsx_file, data_only=True)

print(f"Available sheets: {workbook.sheetnames}\n")

# Inspect each sheet
for sheet_name in workbook.sheetnames:
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print('='*60)

    sheet = workbook[sheet_name]

    # Get dimensions
    print(f"Dimensions: {sheet.dimensions}")
    print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")

    # Show first 5 rows
    print("\nFirst 5 rows:")
    for i, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), 1):
        print(f"Row {i}: {row}")

    # Get header row
    if sheet.max_row > 0:
        print("\nHeader row (Row 1):")
        headers = [cell.value for cell in sheet[1]]
        for i, header in enumerate(headers, 1):
            if header:
                print(f"  Column {i}: '{header}'")

print("\n" + "="*60)
